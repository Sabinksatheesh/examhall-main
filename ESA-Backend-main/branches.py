import openpyxl
from openpyxl import Workbook
import sys
import json

# ACCEPT JSON DATA FROM COMMAND-LINE ARGUMENTS
data_json = sys.argv[1]
data = json.loads(data_json)

for sem in data:
    # LOAD MAIN EXCEL SHEET
    wb = openpyxl.load_workbook('./uploadedExcels/' + sem)
    sheetname = wb.sheetnames
    ws = wb[sheetname[0]]

    # IDENTIFY BRANCHES
    branches = set(a.value for a in ws['D'] if a.value not in (None, 'Branch Name'))
    branch_list = list(branches)
    print("Branches found:", branch_list)

    code_list = []
    sub_list = []

    for sub in branch_list:
        wb_branch = Workbook()
        ws_branchMain = wb_branch.active
        ws_branchMain.title = 'Main'
        r = 1

        print('\nProcessing branch:', sub)
        for p in range(1, ws.max_row + 1):
            if ws.cell(row=p, column=4).value == sub:
                nm = ws.cell(row=p, column=1).value
                if nm is None:
                    continue
                regno = nm[-11:-1]
                email = ws.cell(row=p, column=9).value  # Email from column I
                slot = ws.cell(row=p, column=7).value
                subcode = ws.cell(row=p, column=8).value
                subcode = subcode[-9:-3] if subcode else ""

                ws_branchMain.cell(row=r, column=1).value = nm
                ws_branchMain.cell(row=r, column=2).value = regno
                ws_branchMain.cell(row=r, column=3).value = sub
                ws_branchMain.cell(row=r, column=4).value = slot
                ws_branchMain.cell(row=r, column=5).value = subcode
                ws_branchMain.cell(row=r, column=6).value = email
                r += 1

        if r == 1:
            continue  # Skip empty branch

        codeno = regno[5:7]
        code_list.append(codeno)
        xl_path = './updatedExcels/' + sem[:2] + '_' + codeno + '.xlsx'
        wb_branch.save(xl_path)

        # SORTING ALPHABETICALLY BY REGNO
        sorted_data = sorted(
            list(ws_branchMain.iter_rows(min_row=1, values_only=True)),
            key=lambda x: x[1]  # regno
        )

        wb_branch.remove(ws_branchMain)
        ws_branchSorted = wb_branch.create_sheet('Main')
        for row in sorted_data:
            ws_branchSorted.append(row)

        # IDENTIFY SLOTS
        slot_set = set(a[3] for a in sorted_data if a[3])
        slot_list = sorted(list(slot_set))
        print("Slots found:", slot_list)

        thisdict = {}

        for slot in slot_list:
            wb_branch.create_sheet(slot)
            ws_regular = wb_branch[slot]

            # Determine years
            year_set = set()
            for row in sorted_data:
                if row[3] == slot:
                    regno = row[1]
                    year_set.add(regno[3:5])

            year_list = sorted(list(year_set))
            wb_supply = None
            if len(year_list) != 1:
                wb_supply = wb_branch.create_sheet(slot + '_supply')

            r_reg = 1
            r_sup = 1
            for row in sorted_data:
                if row[3] == slot:
                    regno_year = row[1][3:5]
                    if regno_year == year_list[-1]:
                        for c in range(6):
                            ws_regular.cell(row=r_reg, column=c + 1).value = row[c]
                        r_reg += 1
                    elif wb_supply:
                        for c in range(6):
                            wb_supply.cell(row=r_sup, column=c + 1).value = row[c]
                        r_sup += 1

            print(f'{slot}: Regular = {ws_regular.max_row}', end="")
            thisdict[codeno + slot] = ws_regular.max_row
            if wb_supply and r_sup > 1:
                print(f', Supply = {wb_supply.max_row}')
            else:
                print()

        wb_branch.save(xl_path)
        sub_list.append(thisdict)
