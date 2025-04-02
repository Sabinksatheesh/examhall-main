import openpyxl
import sys
import json

# Get input from command-line argument
data_json = sys.argv[1]
# print("Received JSON input:", data_json)  # Debug

data = json.loads(data_json)

details = data["details"]
sem = details[0].get("sem")
# print("Semester:", sem)  # Debug

s = set()
for item in details:
    slot = item.get("slot")
    # print("Found slot:", slot)  # Debug
    s.add(slot)

slot_list = list(s)
# print("Unique slots:", slot_list)  # Debug

noOfStudents = 0

for input_slot in slot_list:
    code_list = [d["branch"] for d in details if d["slot"] == input_slot]
    # print(f"Processing slot: {input_slot} with branches: {code_list}")  # Debug
    for code in code_list:
        check_supply = 1
        file_path = f'./updatedExcels/S{sem}_{code}.xlsx'
        # print(f"Opening file: {file_path}")  # Debug
        try:
            wb_branch = openpyxl.load_workbook(file_path)
            ws_branch_reg = wb_branch[input_slot]

            # Count normal students
            count_normal = ws_branch_reg.max_row
            # print(f"  Normal students in {input_slot}: {count_normal}")  # Debug
            noOfStudents += count_normal

            # Count supply students if sheet exists
            try:
                ws_branch_sply = wb_branch[f"{input_slot}_supply"]
                count_supply = ws_branch_sply.max_row
                # print(f"  Supply students in {input_slot}_supply: {count_supply}")  # Debug
                noOfStudents += count_supply
            except KeyError:
                # print(f"  No supply sheet for {input_slot}")  # Debug
                check_supply = 0
        except FileNotFoundError:
            # print(f"  File not found: {file_path}")  # Debug
            continue  # Skip missing files

# Room calculations
# print(f"Total students: {noOfStudents}")  # Debug

rem = noOfStudents % 30
noOfRooms30 = noOfStudents // 30
noOfRooms60 = 0

if rem <= 10:
    noOfRooms30 -= 1
    noOfRooms60 = 1
elif rem >= 20:
    noOfRooms30 += 1
else:
    noOfRooms30 -= 2
    noOfRooms60 = 2

# print("Final Room Calculation:")  # Debug
# print("  Rooms of 30 capacity:", noOfRooms30)
# print("  Rooms of 60 capacity:", noOfRooms60)

# Final JSON Output (optional)
# output_data = {
#     "noOfStudents": noOfStudents,
#     "noOfRooms30": noOfRooms30,
#     "noOfRooms60": noOfRooms60
# }

print(noOfStudents)
